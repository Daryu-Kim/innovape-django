import os
import time
import hmac
import hashlib
import math
import random
from django.db.models import Q
from django.http import JsonResponse
from dashboard.models import Product, ProductOptions
from decouple import config
import pandas as pd
from django.conf import settings
import openpyxl as op


def coupang_product_upload(product_codes):
  try:
    # 템플릿 파일 경로
    template_path = os.path.join(settings.MEDIA_ROOT, 'document_forms', 'coupang_upload_form.xlsm')
    
    # 결과 파일 저장할 경로
    timestamp = time.strftime('%Y%m%d_%H%M%S')
    result_dir = os.path.join(settings.MEDIA_ROOT, 'upload_forms')
    os.makedirs(result_dir, exist_ok=True)
    
    # 엑셀 파일 복사 및 새 파일 생성
    wb = op.load_workbook(template_path, keep_vba=True)  # keep_vba=True로 매크로 보존
    ws = wb.active  # 활성 시트 선택
    
    # 2행을 헤더로 설정
    ws.row_dimensions[2].hidden = False  # 2행이 숨겨져 있다면 보이게 설정
    
    # 5행부터 데이터 입력 시작
    current_row = 5
    
    for product_code in product_codes:
      try:
        product = Product.objects.get(product_code=product_code)
        product_options = ProductOptions.objects.filter(product=product)
        
        for index, product_option in enumerate(product_options):
          if "빠른 출고" in product_option.product_option_display_name or "빠른출고" in product_option.product_option_display_name:
            continue
          
          # 무작위 가격 인상률 계산 (15 ~ 40%)
          price_increase_rate = random.uniform(1.15, 1.40)
          increased_price = int(product.product_consumer_price * price_increase_rate)
          consumer_price = math.ceil(increased_price / 100) * 100
          
          # 카테고리 입력
          if product.product_category.filter(category_code__in=['43', '51', '50']).exists():
            ws.cell(row=current_row, column=1, value="[63621] 생활용품>성인용품(19)>흡연용품(19)>전자담배>액상형")  # A열
          elif product.product_category.filter(category_code__in=['52', '45']).exists():
            ws.cell(row=current_row, column=1, value="[63621] 생활용품>성인용품(19)>흡연용품(19)>전자담배>액상형")  # A열
          elif product.product_category.filter(category_code__in=['124', '156']).exists():
            ws.cell(row=current_row, column=1, value="[63624] 생활용품>성인용품(19)>흡연용품(19)>코일/액세서리>카토마이저")  # A열
          elif product.product_category.filter(category_code__in=['157', '158']).exists():
            ws.cell(row=current_row, column=1, value="[63625] 생활용품>성인용품(19)>흡연용품(19)>코일/액세서리>코일")  # A열
          elif product.product_category.filter(category_code__in=['125']).exists():
            ws.cell(row=current_row, column=1, value="[63621] 생활용품>성인용품(19)>흡연용품(19)>전자담배>액상형")  # A열
            
          # 등록상품명 입력
          ws.cell(row=current_row, column=2, value=product.product_seo_title.replace("[", "").replace("]", ""))  # B열
          
          # 상품상태 입력
          ws.cell(row=current_row, column=5, value="새상품")  # E열
          
          # 상태설명 입력
          ws.cell(row=current_row, column=6, value=product.product_description)  # F열
          
          # 브랜드 입력
          ws.cell(row=current_row, column=7, value=product.product_name.split("[")[1].split("]")[0])  # G열
          
          # 제조사 입력
          ws.cell(row=current_row, column=8, value=product.product_name.split("[")[1].split("]")[0])  # H열
          
          # 검색어 입력
          ws.cell(row=current_row, column=9, value=product.product_coupang_keywords.replace(",", "/"))  # I열
          
          # 구매 옵션유형1 입력
          ws.cell(row=current_row, column=10, value="수량")  # J열
          
          # 구매 옵션값1 입력
          if product.product_category.filter(category_code__in=['157', '158']).exists(): 
            ws.cell(row=current_row, column=11, value="1세트")  # K열
          else: 
            ws.cell(row=current_row, column=11, value="1개")  # K열
            
          # 구매 옵션유형2 입력
          ws.cell(row=current_row, column=12, value="색상")  # L열
          
          # 구매 옵션값2 입력
          ws.cell(row=current_row, column=13, value=product_option.product_option_name)  # M열
          
          # 판매가격 입력
          sell_price = product.product_sell_price + product_option.product_option_price if product_option.product_option_price > 0 else product.product_sell_price
          ws.cell(row=current_row, column=62, value=sell_price)  # BJ열
          
          # 할인율기준가 입력
          ws.cell(row=current_row, column=64, value=consumer_price)  # BK열
          
          # 재고수량 입력
          ws.cell(row=current_row, column=65, value=product_option.product_option_stock)  # BL열
          
          # 출고리드타임 입력
          ws.cell(row=current_row, column=66, value=4)  # BM열
          
          # 성인상품 입력
          ws.cell(row=current_row, column=69, value="Y")  # BQ열
          
          # 업체상품코드 입력
          ws.cell(row=current_row, column=73, value=f"{product_code}/{product_option.product_option_code}")  # BU열
          
          # 바코드
          ws.cell(row=current_row, column=75, value="[바코드없음]제조사에서 바코드를 제공 받지 못함")
          
          # 상품고시정보 카테고리 입력
          ws.cell(row=current_row, column=89, value="기타 재화")  # CK열
          
          # 상품고시정보값1 입력
          ws.cell(row=current_row, column=90, value="상품 상세페이지 참조")  # CL열
          
          # 상품고시정보값2 입력
          ws.cell(row=current_row, column=91, value="상품 상세페이지 참조")  # CM열
          
          # 상품고시정보값3 입력
          ws.cell(row=current_row, column=92, value="상품 상세페이지 참조")  # CN열
          
          # 상품고시정보값4 입력
          ws.cell(row=current_row, column=93, value="상품 상세페이지 참조")  # CO열
          
          # 상품고시정보값5 입력
          ws.cell(row=current_row, column=94, value="상품 상세페이지 참조")  # CP열
          
          # 상품 대표이미지 파일명 입력
          thumbnail_filename = product.product_thumbnail_image.url.split("/")[-1]  # 원래 파일명
          thumbnail_name_without_extension = ".".join(thumbnail_filename.split(".")[:-1])  # 확장자 제거
          fixed_thumbnail_filename = f"{thumbnail_name_without_extension}.jpg"  # 확장자를 jpg로 고정
          ws.cell(row=current_row, column=104, value=fixed_thumbnail_filename)  # CZ열
          
          # 상품 상세 설명 파일명 입력
          detail_images = [image.replace('product_detail_images/', '') for image in product.product_detail]
          fixed_detail_images = [f"{'.'.join(img.split('.')[:-1])}.jpg" for img in detail_images]
          detail_images_string = ','.join(fixed_detail_images)
          ws.cell(row=current_row, column=110, value=detail_images_string)  # DF열
          
          if index != (len(product_options) - 1):
            current_row += 1
        
      except Exception as e:
        print(f"Error processing product {product_code}: {e}")
        continue
    
    # 결과 파일 저장
    result_excel = os.path.join(result_dir, f'coupang_upload_result_{timestamp}.xlsm')
    
    wb.save(result_excel)
    return f'coupang_upload_result_{timestamp}.xlsm'
    
  except Exception as e:
    print(e)
    return None