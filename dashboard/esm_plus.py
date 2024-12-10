import os
import time
import math
import random
from django.conf import settings
from dashboard.models import Product, ProductOptions
import openpyxl as op

def esm_plus_product_upload_excel(product_codes):
    try:
        # 500개씩 나누기
        chunk_size = 500
        product_chunks = [product_codes[i:i + chunk_size] for i in range(0, len(product_codes), chunk_size)]
        
        result_files = []
        
        for chunk_index, product_chunk in enumerate(product_chunks):
            # 템플릿 파일 경로
            template_path = os.path.join(settings.MEDIA_ROOT, 'document_forms', 'esm_plus_upload_form.xlsx')
            
            # 결과 파일 저장할 경로
            timestamp = time.strftime('%Y%m%d_%H%M%S')
            result_dir = os.path.join(settings.MEDIA_ROOT, 'upload_forms')
            os.makedirs(result_dir, exist_ok=True)
            
            # 파일명에 청크 번호 추가
            result_filename = f'esm_plus_upload_{timestamp}_part{chunk_index + 1}.xlsx'
            result_path = os.path.join(result_dir, result_filename)

            # 엑셀 파일 복사 및 새 파일 생성
            wb = op.load_workbook(template_path)
            ws = wb.active
            
            # 8행부터 데이터 입력 시작
            current_row = 8
            
            for product_code in product_chunk:
                try:
                    product = Product.objects.get(product_code=product_code)
                    product_options = ProductOptions.objects.filter(product=product)

                    # 무작위 가격 인상률 계산 (15 ~ 40%)
                    price_increase_rate = random.uniform(1.15, 1.40)
                    increased_price = int(product.product_consumer_price * price_increase_rate)
                    consumer_price = math.ceil(increased_price / 100) * 100

                    # 계정 선택
                    ## 노출 사이트
                    ws.cell(row=current_row, column=2, value="옥션/G마켓")  # B열
                    ## 옥션 ID
                    ws.cell(row=current_row, column=3, value="innobite02")  # C열
                    ## G마켓 ID
                    ws.cell(row=current_row, column=4, value="innobite02")  # D열

                    # 상품명
                    ## 상품명
                    ws.cell(row=current_row, column=5, value=product.product_seo_title.replace("[", "").replace("]", ""))  # E열

                    # 카테고리
                    ## 카테고리 코드
                    ## 옥션 노출코드
                    ## G마켓 노출코드
                    if product.product_category.filter(category_code__in=['43', '51', '50', '52', '45']).exists():
                        ws.cell(row=current_row, column=10, value="6304")  # J열
                    elif product.product_category.filter(category_code__in=['124', '156', '157', '158', '125']).exists():
                        ws.cell(row=current_row, column=10, value="6305")  # J열

                    # 판매기간
                    ws.cell(row=current_row, column=14, value="무제한")  # N열

                    # 판매가
                    ## 옥션 판매가
                    ws.cell(row=current_row, column=15, value=consumer_price)  # O열
                    ## G마켓 판매가
                    ws.cell(row=current_row, column=16, value=consumer_price)  # P열

                    # 할인
                    ## 옥션 할인유형
                    ws.cell(row=current_row, column=17, value="정액(원)")  # Q열
                    ## 옥션 할인금액
                    ws.cell(row=current_row, column=18, value=consumer_price-product.product_sell_price)  # R열
                    ## G마켓 할인유형
                    ws.cell(row=current_row, column=19, value="정액(원)")  # S열
                    ## G마켓 할인금액
                    ws.cell(row=current_row, column=20, value=consumer_price-product.product_sell_price)  # T열

                    # 재고수량
                    total_stock = sum(option.product_option_stock for option in product_options)
                    ## 옥션 재고수량
                    ws.cell(row=current_row, column=21, value=total_stock)  # U열
                    ## G마켓 재고수량
                    ws.cell(row=current_row, column=22, value=total_stock)  # V열

                    # 옵션
                    first_option = product_options.first()
                    ## 옵션 타입
                    ws.cell(row=current_row, column=23, value="2개조합형")  # W열
                    ## 옵션명
                    ws.cell(row=current_row, column=24, value=f"출고방식 선택,{first_option.product_option_title}")  # X열
                    ## 옵션 입력값
                    option_values = []
                    for product_option in product_options:
                        option_display = product_option.product_option_display_name.split('/')
                        option_line = f"{option_display[0]},{option_display[1]},정상,노출,{product_option.product_option_stock},{product_option.product_option_stock}"
                        option_values.append(option_line)
                    
                    # 모든 옵션을 줄바꿈으로 연결
                    final_option_value = "\n".join(option_values)
                    ws.cell(row=current_row, column=25, value=final_option_value)  # Y열

                    # 상품이미지
                    ## 기본이미지
                    ws.cell(row=current_row, column=26, value=product.product_origin_thumbnail_image)  # Z열
                    ## 상품상세설명
                    image_urls = product.product_origin_detail
                    html_description = "<div>"

                    for url in image_urls:
                        html_description += f'<img src="{url}" style="width: 100%;">'

                    html_description += "</div>"

                    ws.cell(row=current_row, column=28, value=html_description)  # AB열

                    # 배송정보
                    ## 배송정보 템플릿
                    ws.cell(row=current_row, column=29, value="20417")  # AC열

                    # 상품고시정보
                    ## 상품군 코드
                    ws.cell(row=current_row, column=38, value="35")  # AL열
                    ## 상품고시정보 템플릿 코드
                    ws.cell(row=current_row, column=39, value="222875")  # AM열

                    # 추가정보
                    ## 청소년구매불가여부
                    ws.cell(row=current_row, column=62, value="구매불가")  # BJ열

                    current_row += 1
                except Product.DoesNotExist:
                    continue
            
            # 파일 저장
            wb.save(result_path)
            result_files.append(result_path)
            
        return result_files
        
    except Exception as e:
        raise Exception(f"엑셀 파일 생성 중 오류 발생: {str(e)}")

